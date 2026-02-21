#!/usr/bin/env python3
"""Convert all Excel data files to a single data.js for the dashboard."""
import json
import os
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), "Data final")
OUT_FILE = os.path.join(os.path.dirname(__file__), "data.js")


def read_xlsx(filename, sheet=None):
    """Read an xlsx file and return a list of dicts (one per row)."""
    path = os.path.join(DATA_DIR, filename)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = [str(h).strip() for h in rows[0]]
    data = []
    for row in rows[1:]:
        record = {}
        for h, v in zip(headers, row):
            if v is None:
                record[h] = None
            elif isinstance(v, float):
                record[h] = round(v, 4)
            else:
                record[h] = v
            # Normalise year to int
            if h.lower() in ("año", "anio") and v is not None:
                record[h] = int(v)
        data.append(record)
    return data


datasets = {
    "scorecards": read_xlsx("scorecards_indicadores.xlsx"),
    "seriesHistoricas": read_xlsx("series_historicas_indicadores.xlsx"),
    "pobrezaPanel": read_xlsx("Pobreza_tableau.xlsx", "Data"),
    "pobrezaProvincial": read_xlsx("pobreza_provincial.xlsx"),
    "pobrezaSexoEtnia": read_xlsx("pobreza_sexo_etnia.xlsx"),
    "indicadoresSexoEtnia": read_xlsx("indicadores_sexo_etnia.xlsx"),
    "giniPanel": read_xlsx("gini_panel_tableau.xlsx", "Data"),
    "widIngresoEc": read_xlsx("WID_ingreso_percentiles_tableau.xlsx"),
    "widRiquezaEc": read_xlsx("WID_riqueza_percentiles_tableau.xlsx"),
    "widIngresoALC": read_xlsx("WID_ingreso_percentiles_ALC_tableau.xlsx"),
    "widRiquezaALC": read_xlsx("WID_riqueza_percentiles_ALC_tableau.xlsx"),
}

# Deduplicate WID datasets (they have duplicate rows)
for key in ["widIngresoEc", "widRiquezaEc", "widIngresoALC", "widRiquezaALC"]:
    seen = set()
    unique = []
    for row in datasets[key]:
        frozen = json.dumps(row, sort_keys=True)
        if frozen not in seen:
            seen.add(frozen)
            unique.append(row)
    datasets[key] = unique

with open(OUT_FILE, "w", encoding="utf-8") as f:
    f.write("// Auto-generated — do not edit\n")
    f.write("const DATA = ")
    json.dump(datasets, f, ensure_ascii=False, indent=None)
    f.write(";\n")

print(f"✓ Wrote {OUT_FILE}")
for k, v in datasets.items():
    print(f"  {k}: {len(v)} rows")
