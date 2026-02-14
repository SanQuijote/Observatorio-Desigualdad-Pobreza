"""
create_poverty_tableau.py
--------------------------
Extracts poverty data from INEC tabulados and creates a Tableau-ready Excel file
with poverty indicators from 2007 onwards.

Output: Pobreza_tableau.xlsx
Columns: Año, Mes, Indicador, Nivel, Valor

Indicators:
- Pobreza (income poverty rate)
- Pobreza Extrema (extreme poverty rate)
- NBI (Unsatisfied Basic Needs)
- Pobreza Multidimensional (multidimensional poverty)
- Pobreza Extrema Multidimensional

Levels: Nacional, Urbano, Rural

Usage:
    python3 create_poverty_tableau.py
"""

from pathlib import Path
import pandas as pd
import openpyxl

# Paths
DATA_DIR = Path(__file__).resolve().parent
POVERTY_FILE = DATA_DIR / "pobreza_ingresos_raw" / "202512_Tabulados_Pobreza_EXCEL.XLSX"
NBI_FILE = DATA_DIR / "Tabulados_NBI-dic25.xlsx"
IPM_FILE = DATA_DIR / "Tabulados_IPM-dic25.xlsx"
OUTPUT_FILE = DATA_DIR / "Pobreza_tableau.xlsx"


def extract_poverty_income():
    """Extract income poverty and extreme poverty data (2007-2025)."""
    wb = openpyxl.load_workbook(POVERTY_FILE, data_only=True)
    
    records = []
    
    # Map sheet names to indicator and level
    sheets = [
        ('1.1.1.pobre_nacional', 'Pobreza', 'Nacional'),
        ('1.1.2.pobre_urbana', 'Pobreza', 'Urbano'),
        ('1.1.3.pobre_rural', 'Pobreza', 'Rural'),
        ('1.2.1.extpob_nacional', 'Pobreza Extrema', 'Nacional'),
        ('1.2.2.extpob_urbana', 'Pobreza Extrema', 'Urbano'),
        ('1.2.3.extpob_rural', 'Pobreza Extrema', 'Rural'),
    ]
    
    for sheet_name, indicator, level in sheets:
        ws = wb[sheet_name]
        
        # Find data rows - they start after the header row with "Período"
        # Junio data starts around row 10, Diciembre around row 28
        in_junio = False
        in_diciembre = False
        
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            
            # Check for section headers
            if vals[1] == 'Junio':
                in_junio = True
                in_diciembre = False
                # Don't continue - check if this row also has data
            elif vals[1] == 'Diciembre':
                in_junio = False
                in_diciembre = True
                # Don't continue - check if this row also has data (2007 is on this row!)
            
            # Skip if we haven't reached data yet
            if not in_junio and not in_diciembre:
                continue
            
            # Extract year and value (Incidencia column)
            year = vals[2]
            value = vals[3]
            
            # Stop if we hit notes/footer
            if vals[1] and isinstance(vals[1], str) and 'Fuente' in vals[1]:
                break
            
            # Skip if value is missing/dash
            if value is None or value == '-':
                continue
            
            # Parse year - handle both int and string formats like '    2007 (2)'
            if year is None:
                continue
            
            try:
                # If it's already an int, use it
                if isinstance(year, int):
                    year_int = year
                else:
                    # If it's a string, extract the numeric part
                    year_str = str(year).strip()
                    # Extract first 4-digit number
                    import re
                    match = re.search(r'\d{4}', year_str)
                    if match:
                        year_int = int(match.group())
                    else:
                        continue
                
                value_float = float(value)
            except (ValueError, TypeError):
                continue
            
            # Determine month
            mes = 'Junio' if in_junio else 'Diciembre'
            
            records.append({
                'Año': year_int,
                'Mes': mes,
                'Indicador': indicator,
                'Nivel': level,
                'Valor': value_float
            })
    
    wb.close()
    return pd.DataFrame(records)


def extract_nbi():
    """Extract NBI poverty data (2008-2025)."""
    wb = openpyxl.load_workbook(NBI_FILE, data_only=True)
    
    records = []
    
    sheets = [
        (' 1.1 NBI_nacional', 'Nacional'),
        (' 1.2 NBI_urbano', 'Urbano'),
        (' 1.3 NBI_rural', 'Rural'),
    ]
    
    for sheet_name, level in sheets:
        ws = wb[sheet_name]
        
        # Data starts around row 5, after "Periodo" header
        in_data = False
        
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            
            # Check for data section
            if vals[1] == 'Diciembre':
                in_data = True
                continue
            
            if not in_data:
                continue
            
            # Extract year and value
            year = vals[2]
            value = vals[3]
            
            # Stop at footer
            if vals[1] and isinstance(vals[1], str) and 'Fuente' in vals[1]:
                break
            
            if year is None or value is None:
                continue
            
            try:
                year_int = int(year)
                value_float = float(value)
            except (ValueError, TypeError):
                continue
            
            records.append({
                'Año': year_int,
                'Mes': 'Diciembre',
                'Indicador': 'NBI',
                'Nivel': level,
                'Valor': value_float
            })
    
    wb.close()
    return pd.DataFrame(records)


def extract_multidimensional():
    """Extract multidimensional poverty data (2009-2025)."""
    wb = openpyxl.load_workbook(IPM_FILE, data_only=True)
    ws = wb['1.1. Serie_componentes_IPM']
    
    records = []
    
    # Data starts at row 6 (0-indexed row 5)
    # Columns: Desagregación, Periodo, TPEM (extreme), TPM (total), Intensidad
    # We want TPEM (col 4, point estimate) and TPM (col 7, point estimate)
    
    in_data = False
    current_level = None
    
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        
        # Check for level headers
        if vals[1] in ['Nacional', 'Urbano', 'Rural']:
            current_level = vals[1]
            in_data = True
            continue
        
        if not in_data or current_level is None:
            continue
        
        # Extract year and values
        year = vals[2]
        tpem = vals[4]  # Extreme multidimensional poverty (point estimate)
        tpm = vals[7]   # Total multidimensional poverty (point estimate)
        
        if year is None:
            continue
        
        try:
            year_int = int(year)
        except (ValueError, TypeError):
            continue
        
        # Add extreme multidimensional poverty
        if tpem is not None:
            try:
                records.append({
                    'Año': year_int,
                    'Mes': 'Diciembre',
                    'Indicador': 'Pobreza Extrema Multidimensional',
                    'Nivel': current_level,
                    'Valor': float(tpem)
                })
            except (ValueError, TypeError):
                pass
        
        # Add total multidimensional poverty
        if tpm is not None:
            try:
                records.append({
                    'Año': year_int,
                    'Mes': 'Diciembre',
                    'Indicador': 'Pobreza Multidimensional',
                    'Nivel': current_level,
                    'Valor': float(tpm)
                })
            except (ValueError, TypeError):
                pass
    
    wb.close()
    return pd.DataFrame(records)


def main():
    print("=" * 60)
    print("  Creating Poverty Data for Tableau")
    print("=" * 60)
    
    # Extract all data
    print("\n1. Extracting income poverty data...")
    df_income = extract_poverty_income()
    print(f"   → {len(df_income)} records")
    
    print("\n2. Extracting NBI data...")
    df_nbi = extract_nbi()
    print(f"   → {len(df_nbi)} records")
    
    print("\n3. Extracting multidimensional poverty data...")
    df_ipm = extract_multidimensional()
    print(f"   → {len(df_ipm)} records")
    
    # Combine all data
    print("\n4. Combining all data...")
    df_all = pd.concat([df_income, df_nbi, df_ipm], ignore_index=True)
    
    # Sort by year, indicator, level
    df_all.sort_values(['Año', 'Indicador', 'Nivel'], inplace=True)
    df_all.reset_index(drop=True, inplace=True)
    
    # Filter to only Diciembre data for consistency (since NBI and IPM only have Diciembre)
    # User can choose to keep Junio data if needed
    df_diciembre = df_all[df_all['Mes'] == 'Diciembre'].copy()
    df_diciembre.drop(columns=['Mes'], inplace=True)
    
    print(f"   → Total records (Diciembre only): {len(df_diciembre)}")
    print(f"   → Years covered: {df_diciembre['Año'].min()} - {df_diciembre['Año'].max()}")
    print(f"   → Indicators: {sorted(df_diciembre['Indicador'].unique())}")
    print(f"   → Levels: {sorted(df_diciembre['Nivel'].unique())}")
    
    # Save to Excel
    print(f"\n5. Saving to {OUTPUT_FILE.name}...")
    df_diciembre.to_excel(OUTPUT_FILE, index=False, sheet_name='Data')
    
    # Print summary
    print("\n" + "=" * 60)
    print("  ✅ Poverty data file created successfully!")
    print("=" * 60)
    print(f"\nOutput file: {OUTPUT_FILE}")
    print(f"Total records: {len(df_diciembre)}")
    print("\nData structure:")
    print(df_diciembre.head(10).to_string())
    print("\n" + "=" * 60)
    print("  For Tableau:")
    print("  - X axis: Año")
    print("  - Y axis: Valor")
    print("  - Color/Lines: Indicador")
    print("  - Filter: Nivel (Nacional/Urbano/Rural)")
    print("=" * 60)


if __name__ == "__main__":
    main()
